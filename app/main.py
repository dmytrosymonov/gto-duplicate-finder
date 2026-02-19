"""FastAPI application - web UI and API endpoints."""
import asyncio
import io
import time
import uuid
from collections import deque
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font
from jinja2 import Environment, FileSystemLoader

from app.api_client import fetch_cities, fetch_countries, get_stats
from app.config import DEFAULT_RPS, get_api_key, has_saved_api_key, set_api_key
from app.deduplication import DuplicatePair
from app.scanner import ScanProgress, run_error_scan, run_scan

app = FastAPI(title="GTO Hotel Duplicate Finder")

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)))

# Session-scoped scan state: scan_id -> {scan_id, session_id, progress, results, done, error, city_ids, country_id}
_scans: Dict[str, Dict[str, Any]] = {}
# Queue: deque of {scan_id, session_id, city_ids, country_id, rps, scan_type}
_scan_queue: deque = deque()
_current_scan_id: Optional[str] = None
_cancel_requested_scan_id: Optional[str] = None
# History: scan_id -> {session_id, city_ids, country_id, results, done_at}
_history: Dict[str, Dict[str, Any]] = {}
_HISTORY_TTL_SEC = 2 * 60 * 60  # 2 hours
_MAX_SCANS = 50


@app.middleware("http")
async def session_middleware(request: Request, call_next):
    session_id = request.cookies.get("scan_session_id")
    if not session_id:
        session_id = str(uuid.uuid4())
    request.state.session_id = session_id
    response = await call_next(request)
    if not request.cookies.get("scan_session_id"):
        response.set_cookie(key="scan_session_id", value=session_id, max_age=86400 * 30, httponly=False, samesite="lax")
    return response


def _get_session_id(request: Request) -> str:
    return getattr(request.state, "session_id", "") or str(uuid.uuid4())


def _prune_history() -> None:
    global _history
    now = time.time()
    expired = [sid for sid, h in _history.items() if now - (h.get("done_at") or 0) > _HISTORY_TTL_SEC]
    for sid in expired:
        del _history[sid]


def _render(name: str, **kwargs: Any) -> str:
    t = env.get_template(name)
    return t.render(**kwargs)


@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> str:
    return _render("index.html", api_key_set=has_saved_api_key())


@app.post("/api/apikey")
async def api_set_apikey(request: Request) -> Dict[str, str]:
    body = await request.json()
    key = (body.get("apikey") or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail="apikey required")
    set_api_key(key)
    return {"status": "ok"}


@app.get("/api/countries")
async def api_countries(request: Request) -> Dict[str, Any]:
    if not has_saved_api_key():
        raise HTTPException(status_code=401, detail="API key must be saved first")
    rps = float(request.query_params.get("rps", DEFAULT_RPS))
    try:
        data = await fetch_countries(rps=rps)
        items = [{"id": c.get("id"), "name": c.get("name") or ""} for c in data]
        items.sort(key=lambda x: (x["name"] or "").strip().lower())
        return {"data": items}
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))


@app.get("/api/cities")
async def api_cities(request: Request) -> Dict[str, Any]:
    if not has_saved_api_key():
        raise HTTPException(status_code=401, detail="API key must be saved first")
    country_id = request.query_params.get("country_id")
    if not country_id:
        raise HTTPException(status_code=400, detail="country_id required")
    rps = float(request.query_params.get("rps", DEFAULT_RPS))
    try:
        data = await fetch_cities(int(country_id), rps=rps)
        items = [{"id": c.get("id"), "name": c.get("name") or ""} for c in data]
        items.sort(key=lambda x: (x["name"] or "").strip().lower())
        return {"data": items}
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))


def _prune_scans() -> None:
    global _scans
    while len(_scans) >= _MAX_SCANS:
        done_ids = [sid for sid in list(_scans.keys()) if _scans[sid].get("done")]
        remove_id = done_ids[0] if done_ids else next(iter(_scans.keys()))
        del _scans[remove_id]


def _error_results_to_rows(bad: list) -> list:
    """Convert error scan results to table row format (for error-specific display)."""
    return [
        {
            "hotel_name": b.get("name", ""),
            "id1": b.get("hotel_id"),
            "id2": [],
            "stars": b.get("stars", ""),
            "reason": b.get("reason", "Contains 'Error'"),
            "flag_type": "error",
            "confidence_score": None,
        }
        for b in bad
    ]


def _make_check_cancel(scan_id: str):
    def check() -> bool:
        return _cancel_requested_scan_id == scan_id
    return check


async def _queue_worker() -> None:
    global _cancel_requested_scan_id, _current_scan_id, _scans, _history
    while _scan_queue:
        item = _scan_queue.popleft()
        scan_id = item["scan_id"]
        session_id = item["session_id"]
        city_ids = item["city_ids"]
        country_id = item["country_id"]
        rps = item["rps"]
        scan_type = item.get("scan_type", "duplicates")
        progress = ScanProgress()
        _scans[scan_id] = {
            "scan_id": scan_id,
            "session_id": session_id,
            "progress": progress,
            "results": None,
            "done": False,
            "error": None,
            "city_ids": city_ids,
            "country_id": country_id,
            "scan_type": scan_type,
            "status": "running",
            "started_at": time.time(),
        }
        _current_scan_id = scan_id
        check_cancel = _make_check_cancel(scan_id)
        try:
            country_int = int(country_id) if country_id is not None else None
            if scan_type == "errors":
                bad = await run_error_scan(
                    city_ids=city_ids,
                    country_id=country_int,
                    rps=rps,
                    progress=progress,
                    check_cancel=check_cancel,
                )
                rows = _error_results_to_rows(bad)
            else:
                results = await run_scan(
                    city_ids=city_ids,
                    country_id=country_int,
                    rps=rps,
                    progress=progress,
                    check_cancel=check_cancel,
                )
                rows = _pairs_to_rows(results)
            if scan_id in _scans:
                _scans[scan_id]["results"] = rows
                _scans[scan_id]["done"] = True
                if _cancel_requested_scan_id == scan_id:
                    _scans[scan_id]["error"] = "Отменено пользователем"
                _history[scan_id] = {
                    "session_id": session_id,
                    "city_ids": city_ids,
                    "country_id": country_id,
                    "results": rows,
                    "scan_type": scan_type,
                    "done_at": time.time(),
                }
        except Exception as e:
            if scan_id in _scans:
                _scans[scan_id]["error"] = str(e)
                _scans[scan_id]["done"] = True
                _scans[scan_id]["progress"].error = str(e)
        finally:
            if _cancel_requested_scan_id == scan_id:
                _cancel_requested_scan_id = None
            _current_scan_id = None
            if _scan_queue:
                asyncio.create_task(_queue_worker())


@app.post("/api/scan/cancel")
async def api_scan_cancel(request: Request) -> Dict[str, Any]:
    """Request cancellation of the current scan (if it belongs to the session)."""
    global _cancel_requested_scan_id
    session_id = _get_session_id(request)
    if _current_scan_id:
        scan = _scans.get(_current_scan_id)
        if scan and scan.get("session_id") == session_id:
            _cancel_requested_scan_id = _current_scan_id
            return {"status": "cancelling"}
    return {"status": "no_active_scan"}


@app.post("/api/scan")
async def api_scan_start(request: Request) -> Dict[str, Any]:
    global _scans, _scan_queue
    session_id = _get_session_id(request)
    try:
        body = await request.json() or {}
    except Exception:
        body = {}
    city_ids = body.get("city_ids")
    if city_ids and isinstance(city_ids, list):
        city_ids = [int(x) for x in city_ids if x is not None and str(x).strip() != ""]
    elif body.get("city_id") is not None:
        try:
            city_ids = [int(body["city_id"])]
        except (TypeError, ValueError):
            city_ids = []
    else:
        city_ids = []
    if not city_ids:
        raise HTTPException(
            status_code=400,
            detail="city_ids or city_id required (received: %s)" % (list(body.keys()) if body else "empty body"),
        )
    country_id = body.get("country_id")
    rps = float(body.get("rps", DEFAULT_RPS))
    if rps <= 0 or rps > 100:
        rps = DEFAULT_RPS
    scan_type = body.get("scan_type", "duplicates")
    if scan_type not in ("duplicates", "errors"):
        scan_type = "duplicates"

    if not has_saved_api_key():
        raise HTTPException(status_code=401, detail="API key must be saved first")
    if not get_api_key():
        raise HTTPException(status_code=400, detail="API key not configured")

    _prune_scans()
    _prune_history()
    scan_id = str(uuid.uuid4())
    queue_position = len(_scan_queue) + (1 if _current_scan_id else 0)
    _scan_queue.append({
        "scan_id": scan_id,
        "session_id": session_id,
        "city_ids": city_ids,
        "country_id": country_id,
        "rps": rps,
        "scan_type": scan_type,
    })
    progress = ScanProgress()
    queue_position = len(_scan_queue) + (1 if _current_scan_id else 0)
    _scans[scan_id] = {
        "scan_id": scan_id,
        "session_id": session_id,
        "progress": progress,
        "results": None,
        "done": False,
        "error": None,
        "city_ids": city_ids,
        "country_id": country_id,
        "status": "queued",
    }
    if not _current_scan_id:
        asyncio.create_task(_queue_worker())

    return {"scan_id": scan_id, "status": "queued", "queue_position": queue_position or 1}


@app.get("/api/scan/status")
async def api_scan_status(request: Request) -> Dict[str, Any]:
    global _scans, _scan_queue, _current_scan_id
    session_id = _get_session_id(request)
    stats = get_stats()
    scan_id = request.query_params.get("scan_id")
    if not scan_id:
        return {
            "active": False,
            "done": True,
            "hotels_loaded": 0,
            "comparisons_done": 0,
            "flags_found": 0,
            "results": [],
            "error": None,
            "stats": stats,
            "progress_pct": 0,
        }
    scan = _scans.get(scan_id)
    if not scan or scan.get("session_id") != session_id:
        h = _history.get(scan_id)
        if h and h.get("session_id") == session_id:
            return {
                "active": False,
                "done": True,
                "hotels_loaded": 0,
                "comparisons_done": 0,
                "flags_found": len(h.get("results", [])),
                "results": h.get("results", []),
                "result_type": h.get("scan_type", "duplicates"),
                "error": None,
                "stats": stats,
                "progress_pct": 100,
            }
        return {
            "active": False,
            "done": True,
            "hotels_loaded": 0,
            "comparisons_done": 0,
            "flags_found": 0,
            "results": [],
            "error": None,
            "stats": stats,
            "progress_pct": 0,
        }

    p = scan["progress"]
    progress_pct = getattr(p, "progress_pct", 100 if scan.get("done") else 0)
    resp = {
        "active": not scan.get("done"),
        "done": scan.get("done", False),
        "status": scan.get("status", "running"),
        "hotels_loaded": p.hotels_loaded,
        "comparisons_done": p.comparisons_done,
        "flags_found": p.flags_found,
        "results": None,
        "error": scan.get("error") or p.error,
        "stats": stats,
        "progress_pct": progress_pct,
        "started_at": scan.get("started_at"),
    }

    if scan.get("done") and scan.get("results") is not None:
        resp["results"] = scan["results"]
        resp["result_type"] = scan.get("scan_type", "duplicates")

    return resp


async def _resolve_history_cities_labels(
    raw_items: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Resolve city_ids + country_id to 'City1, Country; City2, Country' for all items."""
    rps = DEFAULT_RPS
    country_names: Dict[Any, str] = {}
    city_maps: Dict[Any, Dict[Any, str]] = {}  # country_id -> {city_id -> city_name}
    unique_country_ids = {it.get("country_id") for it in raw_items if it.get("country_id") is not None}
    if unique_country_ids:
        try:
            countries = await fetch_countries(rps=rps)
            for c in countries:
                cid = c.get("id")
                if cid in unique_country_ids or str(cid) in {str(x) for x in unique_country_ids}:
                    country_names[cid] = (c.get("name") or "").strip()
        except Exception:
            pass
        for cid in unique_country_ids:
            try:
                cities_data = await fetch_cities(int(cid), rps=rps)
                city_maps[cid] = {
                    x.get("id"): (x.get("name") or "").strip()
                    for x in cities_data
                    if x.get("id") is not None
                }
            except Exception:
                city_maps[cid] = {}
    result = []
    for it in raw_items:
        city_ids = it.get("city_ids") or []
        country_id = it.get("country_id")
        if not city_ids:
            result.append({**it, "cities_label": ""})
            continue
        country_name = ""
        if country_id is not None:
            country_name = country_names.get(country_id) or country_names.get(str(country_id)) or ""
        id_to_name = {}
        if country_id is not None:
            id_to_name = city_maps.get(country_id) or city_maps.get(str(country_id)) or {}
        parts = []
        for cid in city_ids:
            name = id_to_name.get(cid) or id_to_name.get(str(cid)) or str(cid)
            parts.append(f"{name}, {country_name}" if country_name else name)
        result.append({**it, "cities_label": "; ".join(parts)})
    return result


@app.get("/api/scan/history")
async def api_scan_history(request: Request) -> Dict[str, Any]:
    if not get_api_key():
        raise HTTPException(status_code=401, detail="API key required")
    _prune_history()
    session_id = _get_session_id(request)
    raw_items = [
        {
            "scan_id": sid,
            "city_ids": h.get("city_ids", []),
            "country_id": h.get("country_id"),
            "flags_count": len(h.get("results", [])),
            "done_at": h.get("done_at"),
        }
        for sid, h in _history.items()
        if h.get("session_id") == session_id
    ]
    raw_items.sort(key=lambda x: -(x.get("done_at") or 0))
    items = await _resolve_history_cities_labels(raw_items)
    return {"data": items}


@app.get("/api/scan/result")
async def api_scan_result(request: Request) -> Dict[str, Any]:
    if not get_api_key():
        raise HTTPException(status_code=401, detail="API key required")
    scan_id = request.query_params.get("scan_id")
    if not scan_id:
        raise HTTPException(status_code=400, detail="scan_id required")
    session_id = _get_session_id(request)
    h = _history.get(scan_id)
    if not h or h.get("session_id") != session_id:
        raise HTTPException(status_code=404, detail="Scan not found")
    return {"results": h.get("results", []), "result_type": h.get("scan_type", "duplicates")}


def _union_find_parent(parent: dict, x: int) -> int:
    if parent[x] != x:
        parent[x] = _union_find_parent(parent, parent[x])
    return parent[x]


def _pairs_to_rows(pairs: list) -> list:
    """Merge duplicate pairs into clusters, return table rows."""
    from app.deduplication import HotelRecord
    if not pairs:
        return []
    id_to_hotel = {}
    pair_by_edge = {}
    for p in pairs:
        id_to_hotel[p.hotel1.id] = p.hotel1
        id_to_hotel[p.hotel2.id] = p.hotel2
        a, b = min(p.hotel1.id, p.hotel2.id), max(p.hotel1.id, p.hotel2.id)
        pair_by_edge[(a, b)] = p
    all_ids = list(id_to_hotel.keys())
    parent = {i: i for i in all_ids}
    for p in pairs:
        a, b = p.hotel1.id, p.hotel2.id
        pa, pb = _union_find_parent(parent, a), _union_find_parent(parent, b)
        if pa != pb:
            parent[pa] = pb
    clusters: dict[int, list[int]] = {}
    for i in all_ids:
        root = _union_find_parent(parent, i)
        clusters.setdefault(root, []).append(i)
    rows = []
    for cluster_ids in clusters.values():
        cluster_ids = sorted(set(cluster_ids))
        if len(cluster_ids) < 2:
            continue
        hotels = [id_to_hotel[i] for i in cluster_ids]
        names = list(dict.fromkeys(h.name for h in hotels if h.name))
        addr_parts = list(dict.fromkeys(h.address for h in hotels if h.address))
        reason = "Группа из %d отелей" % len(cluster_ids)
        flag_type = "review"
        score = 0.0
        for i in range(len(cluster_ids)):
            for j in range(i + 1, len(cluster_ids)):
                a, b = cluster_ids[i], cluster_ids[j]
                p = pair_by_edge.get((min(a, b), max(a, b)))
                if p and (p.confidence_score > score or reason == "Группа из %d отелей" % len(cluster_ids)):
                    reason = p.reason
                    flag_type = p.flag_type
                    score = p.confidence_score
        rows.append({
            "hotel_name": " / ".join(names) if names else "",
            "id1": cluster_ids[0],
            "id2": cluster_ids[1:],
            "address": " | ".join(addr_parts) if addr_parts else "",
            "reason": reason,
            "flag_type": flag_type,
            "confidence_score": round(score, 3),
        })
    return rows


@app.post("/api/export/excel")
async def api_export_excel(request: Request) -> StreamingResponse:
    """Export results to Excel file."""
    try:
        body = await request.json() or {}
    except Exception:
        body = {}
    results = body.get("results") or body.get("data") or []
    if not isinstance(results, list):
        raise HTTPException(status_code=400, detail="results array required")
    result_type = body.get("result_type", "duplicates")

    wb = Workbook()
    ws = wb.active
    if result_type == "errors":
        ws.title = "Ошибки в описаниях"
        headers = ["Название отеля", "ID", "Звёздность"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        for row_idx, r in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=r.get("hotel_name") or "")
            ws.cell(row=row_idx, column=2, value=r.get("id1"))
            ws.cell(row=row_idx, column=3, value=r.get("stars") or "")
        filename = "error_descriptions.xlsx"
    else:
        ws.title = "Дубликаты"
        headers = ["Название отеля", "ID 1", "ID 2", "Адрес", "Общий скоринг", "Причина флага"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        for row_idx, r in enumerate(results, 2):
            id2_val = r.get("id2")
            id2_str = ", ".join(str(x) for x in id2_val) if isinstance(id2_val, list) else str(id2_val or "")
            ws.cell(row=row_idx, column=1, value=r.get("hotel_name") or "")
            ws.cell(row=row_idx, column=2, value=r.get("id1"))
            ws.cell(row=row_idx, column=3, value=id2_str)
            ws.cell(row=row_idx, column=4, value=r.get("address") or "")
            score = r.get("confidence_score")
            ws.cell(row=row_idx, column=5, value=round(score, 3) if score is not None else "")
            ws.cell(row=row_idx, column=6, value=r.get("reason") or "")
        filename = "duplicates.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/stats")
async def api_stats() -> Dict[str, Any]:
    return get_stats()
